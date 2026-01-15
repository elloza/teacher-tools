import streamlit as st
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
import io
import re

st.set_page_config(page_title="Studium Excel 2 Actas", page_icon="📝")

st.title('Studium Excel 2 Actas')
st.write("""
Esta herramienta permite cruzar datos entre dos archivos Excel:
- **Excel de Studium (Moodle)**: archivo de origen con las notas
- **Excel de Actas (USAL)**: archivo de destino que se actualizará

El cruce se realiza por nombres y apellidos usando matching inteligente.
""")

def normalize_name(text):
    """Normaliza un nombre para mejorar el matching"""
    if pd.isna(text):
        return ""
    # Convertir a minúsculas
    text = str(text).lower()
    # Eliminar acentos
    text = text.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    text = text.replace('ñ', 'n')
    # Eliminar caracteres especiales y espacios múltiples
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def find_best_match(student_name, studium_names, threshold=70):
    """
    Encuentra el mejor match para un estudiante usando fuzzy matching

    Args:
        student_name: nombre del estudiante de actas
        studium_names: lista de tuplas (nombre_completo, índice)
        threshold: umbral mínimo de similitud (0-100)

    Returns:
        índice del mejor match o None si no hay match suficientemente bueno
    """
    normalized_student = normalize_name(student_name)

    best_score = 0
    best_index = None

    for name, idx in studium_names:
        normalized_name = normalize_name(name)

        # Calcular varios tipos de similitud
        ratio = fuzz.ratio(normalized_student, normalized_name)
        partial_ratio = fuzz.partial_ratio(normalized_student, normalized_name)
        token_sort_ratio = fuzz.token_sort_ratio(normalized_student, normalized_name)

        # Usar el máximo de las tres métricas
        score = max(ratio, partial_ratio, token_sort_ratio)

        if score > best_score:
            best_score = score
            best_index = idx

    return best_index if best_score >= threshold else None

# Upload Excel de Studium (origen)
st.header("1. Excel de Studium (origen)")
uploaded_studium = st.file_uploader(
    "Sube el archivo Excel de Studium/Moodle",
    type="xlsx",
    key="studium"
)

studium_nombre_col = None
studium_apellidos_col = None
studium_data_col = None
df_studium = None

if uploaded_studium is not None:
    df_studium = pd.read_excel(uploaded_studium)
    st.write("Vista previa del Excel de Studium:")
    st.dataframe(df_studium.head(10))

    col1, col2, col3 = st.columns(3)

    with col1:
        studium_nombre_col = st.selectbox(
            "Columna de Nombre",
            df_studium.columns,
            key="studium_nombre"
        )

    with col2:
        studium_apellidos_col = st.selectbox(
            "Columna de Apellido(s)",
            df_studium.columns,
            key="studium_apellidos"
        )

    with col3:
        studium_data_col = st.selectbox(
            "Columna de datos a copiar (ej: Nota numérica)",
            df_studium.columns,
            key="studium_data"
        )

# Upload Excel de Actas (destino)
st.header("2. Excel de Actas (destino)")
uploaded_actas = st.file_uploader(
    "Sube el archivo Excel de Actas USAL",
    type="xlsx",
    key="actas"
)

actas_nombre_col = None
actas_data_col = None
df_actas = None

if uploaded_actas is not None:
    df_actas = pd.read_excel(uploaded_actas)
    st.write("Vista previa del Excel de Actas:")
    st.dataframe(df_actas.head(10))

    col1, col2 = st.columns(2)

    with col1:
        actas_nombre_col = st.selectbox(
            "Columna de Nombre estudiante",
            df_actas.columns,
            key="actas_nombre"
        )

    with col2:
        actas_data_col = st.selectbox(
            "Columna donde copiar los datos (ej: Nota numérica)",
            df_actas.columns,
            key="actas_data"
        )

# Configuración adicional
st.header("3. Configuración")
col1, col2 = st.columns(2)

with col1:
    threshold = st.slider(
        "Umbral de similitud (%)",
        min_value=50,
        max_value=100,
        value=70,
        help="Porcentaje mínimo de similitud para considerar un match válido"
    )

with col2:
    round_decimals = st.number_input(
        "Redondear a decimales",
        min_value=0,
        max_value=4,
        value=1,
        help="Número de decimales para redondear las notas"
    )

# Botón para procesar
if uploaded_studium is not None and uploaded_actas is not None:
    if st.button('Procesar y cruzar datos', type="primary"):
        with st.spinner('Procesando...'):
            # Crear nombres completos de Studium
            df_studium['nombre_completo'] = (
                df_studium[studium_nombre_col].astype(str) + ' ' +
                df_studium[studium_apellidos_col].astype(str)
            )

            # Crear lista de nombres de Studium con sus índices
            studium_names = [
                (row['nombre_completo'], idx)
                for idx, row in df_studium.iterrows()
            ]

            # Cargar el workbook original de actas para mantener el formato
            uploaded_actas.seek(0)  # Volver al inicio del archivo
            wb = load_workbook(uploaded_actas)
            ws = wb.active

            # Encontrar el índice de las columnas en el Excel
            header_row = 1  # Asumimos que la primera fila es el encabezado
            actas_nombre_col_idx = None
            actas_data_col_idx = None

            for idx, cell in enumerate(ws[header_row], 1):
                if cell.value == actas_nombre_col:
                    actas_nombre_col_idx = idx
                if cell.value == actas_data_col:
                    actas_data_col_idx = idx

            # Estadísticas
            total_students = len(df_actas)
            matched = 0
            not_matched = 0
            match_details = []

            # Procesar cada estudiante de actas
            for row_idx, row in df_actas.iterrows():
                student_name = row[actas_nombre_col]

                # Buscar el mejor match
                best_match_idx = find_best_match(student_name, studium_names, threshold)

                if best_match_idx is not None:
                    # Obtener el valor de Studium
                    value = df_studium.loc[best_match_idx, studium_data_col]

                    # Redondear si es numérico
                    if pd.notna(value) and isinstance(value, (int, float)):
                        value = round(float(value), round_decimals)

                    # Escribir en el Excel (row_idx + 2 porque: +1 por ser 1-indexed, +1 por el header)
                    excel_row = row_idx + 2
                    ws.cell(row=excel_row, column=actas_data_col_idx, value=value)

                    matched += 1
                    studium_name = df_studium.loc[best_match_idx, 'nombre_completo']
                    match_details.append({
                        'Actas': student_name,
                        'Studium': studium_name,
                        'Valor': value,
                        'Match': '✅'
                    })
                else:
                    not_matched += 1
                    match_details.append({
                        'Actas': student_name,
                        'Studium': 'NO ENCONTRADO',
                        'Valor': '',
                        'Match': '❌'
                    })

            # Guardar el workbook modificado en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Mostrar resultados
            st.success('¡Procesamiento completado!')

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total estudiantes", total_students)
            with col2:
                st.metric("Coincidencias", matched)
            with col3:
                st.metric("No encontrados", not_matched)

            # Mostrar detalles de los matches
            st.subheader("Detalle de coincidencias")
            df_matches = pd.DataFrame(match_details)

            # Filtros
            filter_option = st.radio(
                "Mostrar:",
                ["Todos", "Solo coincidencias", "Solo no encontrados"],
                horizontal=True
            )

            if filter_option == "Solo coincidencias":
                df_matches = df_matches[df_matches['Match'] == '✅']
            elif filter_option == "Solo no encontrados":
                df_matches = df_matches[df_matches['Match'] == '❌']

            st.dataframe(df_matches, use_container_width=True)

            # Botón de descarga
            st.download_button(
                label='📥 Descargar Excel de Actas actualizado',
                data=output,
                file_name='actas_actualizado.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # Advertencias
            if not_matched > 0:
                st.warning(
                    f"⚠️ {not_matched} estudiante(s) no se encontraron en el Excel de Studium. "
                    "Revisa los nombres manualmente o ajusta el umbral de similitud."
                )

st.markdown("---")
st.info("""
**💡 Consejos:**
- Asegúrate de que los nombres estén escritos de forma similar en ambos archivos
- Si hay muchos no encontrados, prueba a reducir el umbral de similitud
- El formato del Excel de actas se mantiene intacto (colores, fórmulas, etc.)
- Solo se actualiza la columna seleccionada
""")
