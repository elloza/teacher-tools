import streamlit as st
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
import io
import re
from datetime import datetime

st.set_page_config(page_title="Studium a Actas - Teacher Tools", page_icon="📋")

# Sidebar footer
current_year = datetime.now().year
st.sidebar.markdown(
    f"""
    <div style="position: fixed; bottom: 0; left: 0; width: inherit; padding: 1rem; background: linear-gradient(to top, rgba(255,255,255,1) 80%, rgba(255,255,255,0)); text-align: center;">
        <hr style="margin-bottom: 0.5rem;">
        <p style="color: #666; font-size: 0.8rem; margin: 0;">
            Creado con ❤️ por<br>
            <strong>Álvaro Lozano Murciego</strong><br>
            {current_year}
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

# Inicializar session_state para persistir resultados
if 'processing_done' not in st.session_state:
    st.session_state.processing_done = False
if 'output_file' not in st.session_state:
    st.session_state.output_file = None
if 'match_details' not in st.session_state:
    st.session_state.match_details = []
if 'stats' not in st.session_state:
    st.session_state.stats = {'total': 0, 'matched': 0, 'not_matched': 0}

st.title('📋 Studium a Actas')
st.write("""
Esta herramienta permite cruzar datos entre dos archivos Excel:
- **Excel de Studium (Moodle)**: archivo de origen con las notas
- **Excel de Actas (USAL)**: archivo de destino que se actualizará

El cruce se realiza por nombres y apellidos usando matching inteligente.
""")

def normalize_name(text):
    """Normaliza un nombre para mejorar el matching"""
    if pd.isna(text):
        return ""
    # Convertir a minúsculas
    text = str(text).lower()
    # Eliminar acentos
    text = text.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    text = text.replace('ñ', 'n')
    # Eliminar caracteres especiales y espacios múltiples
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def find_best_match(student_name, studium_names, threshold=70):
    """
    Encuentra el mejor match para un estudiante usando fuzzy matching

    Args:
        student_name: nombre del estudiante de actas
        studium_names: lista de tuplas (nombre_completo, índice)
        threshold: umbral mínimo de similitud (0-100)

    Returns:
        índice del mejor match o None si no hay match suficientemente bueno
    """
    normalized_student = normalize_name(student_name)

    best_score = 0
    best_index = None

    for name, idx in studium_names:
        normalized_name = normalize_name(name)

        # Calcular varios tipos de similitud
        ratio = fuzz.ratio(normalized_student, normalized_name)
        partial_ratio = fuzz.partial_ratio(normalized_student, normalized_name)
        token_sort_ratio = fuzz.token_sort_ratio(normalized_student, normalized_name)

        # Usar el máximo de las tres métricas
        score = max(ratio, partial_ratio, token_sort_ratio)

        if score > best_score:
            best_score = score
            best_index = idx

    return best_index if best_score >= threshold else None

# Upload Excel de Studium (origen)
st.header("1. Excel de Studium (origen)")
uploaded_studium = st.file_uploader(
    "Sube el archivo Excel de Studium/Moodle",
    type="xlsx",
    key="studium"
)

studium_nombre_col = None
studium_apellidos_col = None
studium_data_col = None
df_studium = None

if uploaded_studium is not None:
    df_studium = pd.read_excel(uploaded_studium)
    st.write("Vista previa del Excel de Studium:")
    st.dataframe(df_studium.head(10))

    col1, col2, col3 = st.columns(3)

    with col1:
        studium_nombre_col = st.selectbox(
            "Columna de Nombre",
            df_studium.columns,
            key="studium_nombre"
        )

    with col2:
        studium_apellidos_col = st.selectbox(
            "Columna de Apellido(s)",
            df_studium.columns,
            key="studium_apellidos"
        )

    with col3:
        studium_data_col = st.selectbox(
            "Columna de datos a copiar (ej: Nota numérica)",
            df_studium.columns,
            key="studium_data"
        )

# Upload Excel de Actas (destino)
st.header("2. Excel de Actas (destino)")
uploaded_actas = st.file_uploader(
    "Sube el archivo Excel de Actas USAL",
    type="xlsx",
    key="actas"
)

actas_nombre_col = None
actas_data_col = None
df_actas = None

if uploaded_actas is not None:
    df_actas = pd.read_excel(uploaded_actas)
    st.write("Vista previa del Excel de Actas:")
    st.dataframe(df_actas.head(10))

    col1, col2 = st.columns(2)

    with col1:
        actas_nombre_col = st.selectbox(
            "Columna de Nombre estudiante",
            df_actas.columns,
            key="actas_nombre"
        )

    with col2:
        actas_data_col = st.selectbox(
            "Columna donde copiar los datos (ej: Nota numérica)",
            df_actas.columns,
            key="actas_data"
        )

# Configuración adicional
st.header("3. Configuración")
threshold = st.slider(
    "Umbral de similitud (%)",
    min_value=50,
    max_value=100,
    value=70,
    help="Porcentaje mínimo de similitud para considerar un match válido"
)
st.caption("Las notas se redondean automáticamente a 1 decimal en formato numérico.")

# Botón para procesar
if uploaded_studium is not None and uploaded_actas is not None:
    if st.button('Procesar y cruzar datos', type="primary"):
        # Usar spinner con mensaje claro de carga
        with st.spinner('🔄 Procesando datos... Por favor espera.'):
            # Barra de progreso para feedback visual
            progress_bar = st.progress(0, text="Iniciando procesamiento...")

            # Crear nombres completos de Studium
            progress_bar.progress(10, text="Preparando datos de Studium...")
            df_studium['nombre_completo'] = (
                df_studium[studium_nombre_col].astype(str) + ' ' +
                df_studium[studium_apellidos_col].astype(str)
            )

            # Crear lista de nombres de Studium con sus índices
            studium_names = [
                (row['nombre_completo'], idx)
                for idx, row in df_studium.iterrows()
            ]

            # Cargar el workbook original de actas para mantener el formato
            progress_bar.progress(20, text="Cargando Excel de Actas...")
            uploaded_actas.seek(0)  # Volver al inicio del archivo
            wb = load_workbook(uploaded_actas)
            ws = wb.active

            # Encontrar el índice de las columnas en el Excel
            header_row = 1  # Asumimos que la primera fila es el encabezado
            actas_nombre_col_idx = None
            actas_data_col_idx = None

            for idx, cell in enumerate(ws[header_row], 1):
                if cell.value == actas_nombre_col:
                    actas_nombre_col_idx = idx
                if cell.value == actas_data_col:
                    actas_data_col_idx = idx

            # Estadísticas
            total_students = len(df_actas)
            matched = 0
            not_matched = 0
            match_details = []

            progress_bar.progress(30, text="Cruzando datos de estudiantes...")

            # Procesar cada estudiante de actas
            for i, (row_idx, row) in enumerate(df_actas.iterrows()):
                student_name = row[actas_nombre_col]

                # Actualizar progreso
                progress_pct = 30 + int((i / total_students) * 60)
                progress_bar.progress(progress_pct, text=f"Procesando estudiante {i+1} de {total_students}...")

                # Buscar el mejor match
                best_match_idx = find_best_match(student_name, studium_names, threshold)

                # Posición en Excel (row_idx + 2 porque: +1 por ser 1-indexed, +1 por el header)
                excel_row = row_idx + 2

                if best_match_idx is not None:
                    # Obtener el valor de Studium
                    value = df_studium.loc[best_match_idx, studium_data_col]
                    studium_name = df_studium.loc[best_match_idx, 'nombre_completo']

                    # Verificar si el valor está vacío o es NaN
                    if pd.isna(value) or (isinstance(value, str) and value.strip() == ''):
                        # Columna numérica vacía, NP en la columna de la derecha (texto)
                        ws.cell(row=excel_row, column=actas_data_col_idx, value=None)
                        ws.cell(row=excel_row, column=actas_data_col_idx + 1, value="NP")
                        value = "NP"
                    else:
                        # Si es numérico, redondear a 1 decimal y escribir como número
                        try:
                            numeric_value = float(value)
                            value = round(numeric_value, 1)
                            ws.cell(row=excel_row, column=actas_data_col_idx, value=value)
                        except (ValueError, TypeError):
                            # Si no es convertible a número, escribir tal cual
                            ws.cell(row=excel_row, column=actas_data_col_idx, value=value)

                    matched += 1
                    match_details.append({
                        'Actas': student_name,
                        'Studium': studium_name,
                        'Valor': value,
                        'Match': '✅'
                    })
                else:
                    # No hay match -> columna numérica vacía, NP en la columna de la derecha
                    not_matched += 1
                    ws.cell(row=excel_row, column=actas_data_col_idx, value=None)
                    ws.cell(row=excel_row, column=actas_data_col_idx + 1, value="NP")
                    match_details.append({
                        'Actas': student_name,
                        'Studium': 'NO ENCONTRADO',
                        'Valor': 'NP',
                        'Match': '❌'
                    })

            # Guardar el workbook modificado en memoria
            progress_bar.progress(95, text="Generando archivo Excel...")
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            progress_bar.progress(100, text="¡Completado!")

            # Guardar en session_state para persistir después de la recarga
            st.session_state.processing_done = True
            st.session_state.output_file = output.getvalue()
            st.session_state.match_details = match_details
            st.session_state.stats = {
                'total': total_students,
                'matched': matched,
                'not_matched': not_matched
            }

# Mostrar resultados si hay datos procesados (persiste después de descargar)
if st.session_state.processing_done:
    st.success('¡Procesamiento completado!')

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total estudiantes", st.session_state.stats['total'])
    with col2:
        st.metric("Coincidencias", st.session_state.stats['matched'])
    with col3:
        st.metric("No encontrados", st.session_state.stats['not_matched'])

    # Mostrar detalles de los matches
    st.subheader("Detalle de coincidencias")
    df_matches = pd.DataFrame(st.session_state.match_details)

    # Filtros
    filter_option = st.radio(
        "Mostrar:",
        ["Todos", "Solo coincidencias", "Solo no encontrados"],
        horizontal=True,
        key="filter_results"
    )

    df_filtered = df_matches.copy()
    if filter_option == "Solo coincidencias":
        df_filtered = df_filtered[df_filtered['Match'] == '✅']
    elif filter_option == "Solo no encontrados":
        df_filtered = df_filtered[df_filtered['Match'] == '❌']

    st.dataframe(df_filtered, use_container_width=True)

    # Botón de descarga (fuera del bloque del botón para evitar recarga)
    st.download_button(
        label='📥 Descargar Excel de Actas actualizado',
        data=st.session_state.output_file,
        file_name='actas_actualizado.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key="download_result"
    )

    # Advertencias
    if st.session_state.stats['not_matched'] > 0:
        st.warning(
            f"⚠️ {st.session_state.stats['not_matched']} estudiante(s) no se encontraron en el Excel de Studium. "
            "Revisa los nombres manualmente o ajusta el umbral de similitud."
        )

    # Botón para limpiar y procesar de nuevo
    if st.button("🔄 Limpiar y procesar nuevos archivos", type="secondary"):
        st.session_state.processing_done = False
        st.session_state.output_file = None
        st.session_state.match_details = []
        st.session_state.stats = {'total': 0, 'matched': 0, 'not_matched': 0}
        st.rerun()

st.markdown("---")
st.info("""
**💡 Consejos:**
- Asegúrate de que los nombres estén escritos de forma similar en ambos archivos
- Si hay muchos no encontrados, prueba a reducir el umbral de similitud
- El formato del Excel de actas se mantiene intacto (colores, fórmulas, etc.)
- Solo se actualiza la columna seleccionada
""")
